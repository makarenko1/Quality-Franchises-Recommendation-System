
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from recommendations_algorithm import (
    load_movies_metadata,
    load_dialogue_features,
    load_franchise_map,
    load_svd_model,
    recommend,
)
from interactive_baselines import load_rating_stats, recommend_baseline


SURVEY_PATH   = "מחט בערימת דאטה- תוצאות סקר.xlsx"
DATASET_PATH  = "dataset.csv"
MOVIES_32M     = "datasets/movies-32M/movies_clean.csv"
OUTPUT_PATH   = "recommendations_output.xlsx"
RATINGS_PATH = "datasets/movies-32M/movies_ratings_clean.csv"


def _safe_movie_id(val):
    """Return int movie ID if val looks like one (>10), else None."""
    try:
        v = float(val)
        return int(v) if v > 10 else None
    except (TypeError, ValueError):
        return None


def parse_survey(survey_path, id_to_title):
    """
    Parse survey Excel into a list of respondent dicts.
    Handles two row formats:
      - clean  : rank{n}_movie_id / rank{n}_title / rank{n}_year in correct columns
      - shifted: scores were inserted after each movie_id, shifting later columns left

    Returns:list of dict with keys: name, inputs (list of (movie_id, title))
    """
    df = pd.read_excel(survey_path)

    respondents = {}
    for _, row in df.iterrows():
        name = str(row.get("name", "")).strip()
        if not name or name == "nan":
            continue

        # Detect shifted format: rank2_movie_id holds a score (1-10)
        try:
            r2_val = float(row["rank2_movie_id"])
            shifted = 1 <= r2_val <= 10
        except (TypeError, ValueError):
            shifted = False

        if not shifted:
            raw_ids = [
                _safe_movie_id(row["rank1_movie_id"]),
                _safe_movie_id(row["rank2_movie_id"]),
                _safe_movie_id(row["rank3_movie_id"]),
            ]
            raw_titles = [
                str(row["rank1_title"]),
                str(row["rank2_title"]),
                str(row["rank3_title"]),
            ]
        else:
            # Columns shifted: score inserted after rank1_movie_id
            raw_ids = [
                _safe_movie_id(row["rank1_movie_id"]),
                _safe_movie_id(row["rank2_title"]),   # real rank2 id
                _safe_movie_id(row["rank3_year"]),    # real rank3 id
            ]
            raw_titles = [
                str(row["rank1_title"]),
                str(row["rank2_year"]),               # real rank2 title
                str(row["rank1_score"]),              # real rank3 title
            ]

        # Keep only IDs that exist in the dataset
        inputs = [
            (mid, id_to_title.get(mid, title))
            for mid, title in zip(raw_ids, raw_titles)
            if mid and mid in id_to_title
        ]

        # Deduplicate: keep latest submission per name
        respondents[name] = {"name": name, "inputs": inputs}

    return [r for r in respondents.values() if len(r["inputs"]) >= 2]



def get_recommendations(respondents, movies_meta, movie_ids, movie_factors,
                         movie_biases, dq, fm, movies_1m, stats, n=3):
    """
    For each respondent, run all 4 methods and return results.

    Returns
    -------
    list of dict:
        name        : str
        inputs      : list of title strings
        our_algorithm, popular, highest_rated, random : list of "Title (Year)" strings
    """
    def fmt(recs):
        return [
            f"{r['title']} ({r['year']})" if r.get("year") else r["title"]
            for r in recs
        ]

    results = []
    for resp in respondents:
        name         = resp["name"]
        input_titles = [t for _, t in resp["inputs"]]

        print(f"  Processing {name}: {input_titles}")

        alg = fmt(recommend(
            input_titles, movies_meta, movie_ids,
            movie_factors, movie_biases, dq, fm, n=n,
        ))
        pop = fmt(recommend_baseline(input_titles, "popular",       movies_1m, stats, n=n))
        hr  = fmt(recommend_baseline(input_titles, "highest_rated", movies_1m, stats, n=n))
        rnd = fmt(recommend_baseline(input_titles, "random",        movies_1m, stats, n=n))

        results.append({
            "name":          name,
            "inputs":        input_titles,
            "our_algorithm": alg,
            "popular":       pop,
            "highest_rated": hr,
            "random":        rnd,
        })

    return results


# Excel output

METHOD_SECTIONS = [
    ("Our Algorithm",  "our_algorithm", "alg_bg"),
    ("Popular",        "popular",       "pop_bg"),
    ("Highest Rated",  "highest_rated", "hr_bg"),
    ("Random",         "random",        "rnd_bg"),
]


def _thin_border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)



def write_person_sheet(wb, result):
    sheet_name = (
        result["name"][:31]
        .replace("/", "-").replace("\\", "-")
        .replace("*", "").replace("?", "").replace(":", "")
        .replace("[", "").replace("]", "")
    )
    ws = wb.create_sheet(title=sheet_name)
    row = 1

    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row, 1, f"Movie Recommendations for {result['name']}")
    c.font = Font(name="Arial", bold=True, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28
    row += 1

    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row, 1, "Please rate each recommended movie from 1 (poor) to 5 (excellent) in the 'Your Rating' column.")
    c.font = Font(name="Arial", italic=True, size=10)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 18
    row += 2

    ws.merge_cells(f"A{row}:D{row}")
    c = ws.cell(row, 1, "Your Selected Movies")
    c.font = Font(name="Arial", bold=True, size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20
    row += 1

    for movie in result["inputs"]:
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row, 1, f"  ▸  {movie}")
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1

    for label, key, _ in METHOD_SECTIONS:
        ws.merge_cells(f"A{row}:D{row}")
        c = ws.cell(row, 1, label)
        c.font = Font(name="Arial", bold=True, size=11)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        row += 1

        for col, hdr in enumerate(["#", "Recommended Movie", "Your Rating (1–5)", "Notes"], 1):
            c = ws.cell(row, col, hdr)
            c.font = Font(name="Arial", bold=True, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _thin_border()
        ws.row_dimensions[row].height = 18
        row += 1

        for i, movie in enumerate(result[key], 1):
            ws.cell(row, 1, i).font = Font(name="Arial", size=10)
            ws.cell(row, 1).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row, 1).border = _thin_border()

            ws.cell(row, 2, movie).font = Font(name="Arial", size=10)
            ws.cell(row, 2).alignment = Alignment(horizontal="left", vertical="center")
            ws.cell(row, 2).border = _thin_border()

            ws.cell(row, 3, "").border = _thin_border()
            ws.cell(row, 3).alignment = Alignment(horizontal="center", vertical="center")

            ws.cell(row, 4, "").border = _thin_border()
            ws.row_dimensions[row].height = 20
            row += 1

        row += 1

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 25

def write_excel(results, output_path):
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet
    for result in results:
        write_person_sheet(wb, result)
    wb.save(output_path)
    print(f"\nSaved: {output_path}  ({len(results)} sheets)")


def main():
    movies_meta = load_movies_metadata(DATASET_PATH)
    dq          = load_dialogue_features(DATASET_PATH)
    fm          = load_franchise_map(DATASET_PATH)
    movie_ids, movie_factors, movie_biases = load_svd_model(RATINGS_PATH)
    movies_1m   = pd.read_csv(MOVIES_32M, low_memory=False)
    stats       = load_rating_stats(RATINGS_PATH)

    id_to_title = dict(zip(
        movies_meta["MovieID"].astype(int),
        movies_meta["Title"].astype(str),
    ))

    print("\nParsing survey...")
    respondents = parse_survey(SURVEY_PATH, id_to_title)
    print(f"  {len(respondents)} valid respondents found")

    print("\nGenerating recommendations...")
    results = get_recommendations(
        respondents, movies_meta, movie_ids,
        movie_factors, movie_biases, dq, fm, movies_1m, stats, n=3,
    )

    print("\nWriting Excel...")
    write_excel(results, OUTPUT_PATH)


if __name__ == "__main__":
    main()
